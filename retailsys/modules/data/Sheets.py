from config import *
from openpyxl import load_workbook


class Sheet:
    def __init__(self, path):
        self.path = path
        self._wb = load_workbook(path)
        self.sheet = self._wb.active

    def get_columns(self, columns, offset=0):
        data = []
        for i in columns:
            data.append(self.get_column(i, offset))
        return [[data[j][i] for j in range(len(data))] for i in range(len(data[0]))]

    def get_column(self, column, offset=0):
        data = []
        ind = 1 + offset
        while self.get_value(column, ind):
            data.append(self.get_value(column, ind))
            ind += 1
        return data

    def get_value(self, column, line):
        val = self.sheet[f'{column}{line}'].value
        if val:
            return str(val)
        return self.sheet[f'{column}{line}'].value

    def set_value(self, column, line, val):
        self.sheet[f'{column}{line}'].value = val

    def _save(self):
        self._wb.save(self.path)

    def close(self):
        self._save()

    def open(self):
        self._wb = load_workbook(self.path)
        self.sheet = self._wb.active

    def __getitem__(self, item):
        return self.get_value(item, '')

    def __setitem__(self, key, value):
        return self.set_value(key, "", value)


class MainSheet(Sheet):
    def __init__(self):
        super().__init__(PATH_MAIN_SHEET)

    def get_special_data(self, name, line_index):
        if name == "count_line":
            return self.get_value("S", line_index)
        if name == "first_sort_key":
            return self.get_value("P", line_index)
        if name == "second_sort_key":
            return self.get_value("M", line_index)

    def get_sorted_lines(self, sort_keys):
        sorted_lines = []
        config_lines = self.load_config_lines()

        for line in config_lines:
            for number in range(COUNT_GROUPS):
                if line[0] in sort_keys[number]:
                    sorted_lines[number].append(line[1])

        return sorted_lines

    def load_config_lines(self):
        line_index = 2
        config_lines = []
        while self.get_value(LINE_MATCHER["COUNT_LINES"], line_index):
            count_lines = int(self.get_value(LINE_MATCHER["COUNT_LINES"], line_index))
            config_lines.append([
                [self.get_value(LINE_MATCHER["first_sort_key"], line_index),
                 self.get_value(LINE_MATCHER["second_sort_key"], line_index)],
                count_lines]
            )
            line_index += count_lines
        return config_lines


class FilterSheet(Sheet):
    def __init__(self):
        super().__init__(PATH_SORT_SHEET)

    def get_sort_keys(self):
        return [self.get_columns(columns) for columns in SHEET_SORT_COLUMN]


class PlansSheet(Sheet):
    def __init__(self):
        super().__init__(PATH_PLANS_SHEET)
        self.self_plans_set = self.load()

    def get_plans(self):
        return self.self_plans_set

    def load(self):
        out = []
        counter = 1
        while self.get_value(SHEET_PLANS_COLUMN[0], counter):
            line = []
            for column in SHEET_PLANS_COLUMN:
                line.append(self.get_value(column, counter))
            out.append(line)
            counter += 1
        return out


class ErrorsSheetBase(Sheet):
    def __init__(self):
        super().__init__(PATH_ERRORS_SHEET)
        self.errors = []

        index = 2
        while self.get_value("A", index):
            self.errors.append({
                "name": self.get_value("A", index),
                "replace": self.get_value("B", index),
                "flags": self.get_value("C", index),
            })
            index += 1

    def get_error_description(self, name):
        for i in self.errors:
            if i['name'] == name:
                return i
        return {
            'name': 'undefined'
        }


class ErrorsSheetCollect(Sheet):
    def __init__(self):
        super().__init__(PATH_COLLECT_ERRORS_SHEET)
        self.index = 1
        while self.get_value("A", self.index):
            self.index += 1

    def write_error(self, error, data=""):
        self.open()
        self.set_value("A", self.index, error)
        self.set_value("B", self.index, data)
        self.close()
        self.index += 1